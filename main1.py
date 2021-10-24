import os
import tkinter as tk
from tkinter import *

import openpyxl

import test1

count = 0

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
root = tk.Tk()
root.geometry("650x50")
cwd = os.path.expanduser(r'~\Desktop\prova.xlsx')


def atribute():
    global count
    count += 1
    nr1 = test1.shfaqi()
    if count % 2 == 0:
        root.geometry("650x50")
        button_4['text'] = 'SHFAQ'
    else:
        root.geometry("650x500")
        wb = openpyxl.load_workbook(cwd)
        ws = wb.active
        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['B' + str(nr1 - 4)].value))
        a1.configure(state='readonly')
        a1.place(x=10, y=100, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['C' + str(nr1 - 4)].value))
        a1.configure(state='readonly')
        a1.place(x=135, y=100, width=160, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['E' + str(nr1 - 4)].value))
        a1.configure(state='readonly')
        a1.place(x=310, y=100, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['B' + str(nr1 - 3)].value))
        a1.configure(state='readonly')
        a1.place(x=10, y=140, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['C' + str(nr1 - 3)].value))
        a1.configure(state='readonly')
        a1.place(x=135, y=140, width=160, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['E' + str(nr1 - 3)].value))
        a1.configure(state='readonly')
        a1.place(x=310, y=140, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['B' + str(nr1 - 2)].value))
        a1.configure(state='readonly')
        a1.place(x=10, y=180, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['C' + str(nr1 - 2)].value))
        a1.configure(state='readonly')
        a1.place(x=135, y=180, width=160, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['E' + str(nr1 - 2)].value))
        a1.configure(state='readonly')
        a1.place(x=310, y=180, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['B' + str(nr1 - 1)].value))
        a1.configure(state='readonly')
        a1.place(x=10, y=220, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['C' + str(nr1 - 1)].value))
        a1.configure(state='readonly')
        a1.place(x=135, y=220, width=160, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['E' + str(nr1 - 1)].value))
        a1.configure(state='readonly')
        a1.place(x=310, y=220, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['B' + str(nr1)].value))
        a1.configure(state='readonly')
        a1.place(x=10, y=260, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['C' + str(nr1)].value))
        a1.configure(state='readonly')
        a1.place(x=135, y=260, width=160, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['E' + str(nr1)].value))
        a1.configure(state='readonly')
        a1.place(x=310, y=260, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['B' + str(nr1 + 1)].value))
        a1.configure(state='readonly')
        a1.place(x=10, y=300, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['C' + str(nr1 + 1)].value))
        a1.configure(state='readonly')
        a1.place(x=135, y=300, width=160, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['E' + str(nr1 + 1)].value))
        a1.configure(state='readonly')
        a1.place(x=310, y=300, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['B' + str(nr1 + 2)].value))
        a1.configure(state='readonly')
        a1.place(x=10, y=340, width=120, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['C' + str(nr1 + 2)].value))
        a1.configure(state='readonly')
        a1.place(x=135, y=340, width=160, height=30)

        a1 = tk.Entry(root, font="Calibri 12")
        a1.insert(0, str(ws['E' + str(nr1 + 2)].value))
        a1.configure(state='readonly')
        a1.place(x=310, y=340, width=120, height=30)

        button_4['text'] = 'FSHI'


def fillimi():
    test1.main(True)
    button_1['state'] = tk.DISABLED
    button_3['state'] = tk.NORMAL


def stop():
    test1.main(False)
    button_1['state'] = tk.NORMAL
    button_3['state'] = tk.DISABLED


# def procesi(proc):
# e1.insert(0,proc)
eshtegjalle = test1.drita

button_1 = Button(root, text='NIS', command=fillimi, padx=5, pady=5, state=tk.NORMAL)
button_1.place(x=450, y=10)
button_3 = Button(root, text='MBYLL', command=stop, padx=5, pady=5, state=tk.DISABLED)
button_3.place(x=500, y=10)
button_4 = Button(root, text='SHFAQ', command=atribute, padx=5, pady=5, state=tk.NORMAL)
button_4.place(x=570, y=10)
box_value = StringVar()
label = Label(root, text='eshtegjalle')
label.place(x=10, y=10)
# e1 = tk.Entry(root, font="Calibri 15 bold")
# e1.place(x=10, y=10, width=200, height=30)
# root.resizable(2,2)
root.wm_attributes("-toolwindow", 1)
# root.call("wm", "attributes", ".", "-modified", "0.9")
root.wm_attributes("-topmost", 1)
# root.attributes('-alpha',0.5)

root.title("Exceli")
root.mainloop()
