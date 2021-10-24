import getpass
import os
import sched
import threading
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

import app

cwd = os.path.expanduser(r'~\Desktop\prova.xlsx')


def shfaqi():
    wb = load_workbook(cwd)
    ws = wb.active
    for i in range(2, 1000):
        if ws['B' + str(i)].value is not None:
            pass
        elif ws['B' + str(i)].value is None:
            return i


def shfaqe(programi, koha):
    wb = load_workbook(cwd)
    ws = wb.active

    for i in range(2, 1000):
        if ws['B' + str(i)].value is not None:
            pass
        elif ws['B' + str(i)].value is None:
            print(ws['B' + str(i - 1)].value)
            ws['B' + str(i)].value = programi
            ws['C' + str(i)].value = koha
            ws['D' + str(i)].value = ws['C' + str(i - 1)].value
            wb.save(cwd)
            return i


s = sched.scheduler(time.time, time.sleep)


# noinspection PyGlobalUndefined
def perseritarog():
    vlera = 0
    global dead, vlera1
    global testi

    while dead:
        kushti = os.path.exists('\\Users\\' + getpass.getuser() + '\\Desktop\\prova.xlsx')
        print(kushti)
        if kushti:
            vlera1 = (str(app.get_active_window()))
            if vlera1 == vlera:
                # main.procesi(vlera1)
                pass
            else:
                now = datetime.now()
                shfaqe(str(app.get_active_window()), now)
        if not kushti:
            wbi = Workbook()
            wbi.save(cwd)
            vlera1 = (str(app.get_active_window()))
            if vlera1 == vlera:
                # main.procesi(vlera1)
                pass
            else:
                now = datetime.now()
                shfaqe(str(app.get_active_window()), now)

        vlera = vlera1


# perseritarog()
# noinspection PyGlobalUndefined
def main(pro):
    global dead
    dead = pro
    # dead = True
    our_thread = threading.Thread(target=perseritarog)
    our_thread.start()


def drita():
    return 'lalalala'
